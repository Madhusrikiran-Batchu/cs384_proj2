import csv 
from openpyxl.workbook import Workbook
import os
import pandas as pd
import numpy as np
from fpdf import FPDF
from datetime import datetime
import glob
import PySimpleGUI as sg


sg.theme('LightBlue2')
layout = [[sg.Text('Gui for window for Python Project 2')],      
                 [sg.Text('Browse for Stamp seal ')],
                 [sg.Input(key = '-pic-'), sg.FileBrowse()],
                 [sg.Text('Enter the list of roll numbers to be printed')],
                 [sg.Text('From'), sg.InputText(key='-IN-')],
                 [sg.Text('To'), sg.InputText(key='-ID-')],
                 [sg.Button('Generate Roll number transcripts')],
                 [sg.Button('Generate all transcripts')],
                 [sg.Submit(), sg.Cancel()]]    


window = sg.Window('ORIGINAL').Layout(layout)    
while True:             # Event Loop
    event, values = window.Read()
    pic = values['-pic-']
    fro = values['-IN-']
    too = values['-ID-']
    if event in (None, 'Exit'):
        break
    if event == 'Generate all transcripts':
        all = 1
        break
    elif event == 'Generate Roll number transcripts':
        all = 0
        break 
window.close()
sg.popup('PLease wait pdfs are being generated')
print(fro)
print(too)
fro_rol = int(fro[6:])
to_rol = int(too[6:])
print(fro_rol)
print(to_rol)
print(to_rol-fro_rol)
numb = to_rol - fro_rol

grades = {}
names_roll = {}
subjects_master = {}


file_names_roll = open('sample_input/names-roll.csv', 'r')
file_subject_master = open("sample_input/subjects_master.csv", 'r')
file_grades = open("sample_input/grades.csv", 'r')

reader1 = csv.DictReader(file_names_roll)        #read data with reader.
for row in reader1:
  names_roll[row['Roll']] = row['Name']

reader2 = csv.DictReader(file_subject_master)    #read data with reader.
for row in reader2:
  subjects_master[row['subno']] = [row['subname'],row['ltp'],row['crd']]
  
reader3 = csv.DictReader(file_grades)
for row in reader3:
  if not row['Roll'] in grades.keys():
    grades[row['Roll']] = {}
    if row['Sem'] not in grades[row['Roll']].keys():
      grades[row['Roll']][row['Sem']] = [[ row['SubCode'],
                                    subjects_master[row['SubCode']][0],
                                    subjects_master[row['SubCode']][1],
                                    subjects_master[row['SubCode']][2],
                                    row['Sub_Type'],
                                    row['Grade']]]
    else:
       grades[row['Roll']][row['Sem']].append([ row['SubCode'],
                                    subjects_master[row['SubCode']][0],
                                    subjects_master[row['SubCode']][1],
                                    subjects_master[row['SubCode']][2],
                                    row['Sub_Type'],
                                    row['Grade']])
  else:
    if row['Sem'] not in grades[row['Roll']].keys():
      grades[row['Roll']][row['Sem']] = [[ row['SubCode'],
                                    subjects_master[row['SubCode']][0],
                                    subjects_master[row['SubCode']][1],
                                    subjects_master[row['SubCode']][2],
                                    row['Sub_Type'],
                                    row['Grade']]]
    else:
       grades[row['Roll']][row['Sem']].append([ row['SubCode'],
                                    subjects_master[row['SubCode']][0],
                                    subjects_master[row['SubCode']][1],
                                    subjects_master[row['SubCode']][2],
                                    row['Sub_Type'],
                                    row['Grade']])
                                    


grade_dic = {'BB': 8, 'BC': 7, 'AB' : 9, 'CC' : 6, 'AA' : 10, 'CD' : 5, 'DD' : 4, 'F' : 0, 'F*' : 0, 'DD*' : 4, ' BB' : 8}
for roll in grades.keys():
  overall = []
  overall.append(['Roll No.', roll])
  overall.append(['Name of the Student',names_roll[roll]])
  overall.append(['Discipline',roll[4:6]])
  semester = ['Semester No.']
  sem_crd = ['Semester wise Credit Taken']
  spi = ['SPI']
  total_credits = ['Total Credits Taken']
  cpi = ['CPI']

  filepath = 'output/' + roll + '.xlsx'
  directory = os.path.dirname(filepath)

  if not os.path.exists(directory):
    os.makedirs(directory)

  wb = Workbook()
  
  for sem in grades[roll].keys():
    semester.append(sem)
    s_num = 1 
    credit = 0
    s_p_i = 0 
    ws = wb.create_sheet()
    ws.title = 'Sem' + sem
    ws.append(['Sl No.','Subject code','Subject Name','L-T-P','Credit','Subject Type','Grade'])
    
    for data in grades[roll][sem]:
      s_p_i += int(data[3]) * grade_dic[data[5]]
      credit += int(data[3]) 
      data.insert(0,s_num)
      ws.append(data)
      s_num += 1

      
      
    sem_crd.append(credit)
    spi.append(round(s_p_i/credit,2))

    if type(total_credits[-1]) == str:
      total_credits.append(credit)
    else:
      a = credit
      a += total_credits[-1] + credit
      total_credits.append(a)
    

  overall.append(semester)
  overall.append(sem_crd)
  overall.append(spi)
  overall.append(total_credits)
  ws = wb['Sheet']
  for row in overall:
    ws.append(row)
  wb.save(filename=filepath)

class PDF(FPDF):
    def border_lines(self):
        self.set_line_width(0.0)
        self.rect(5,5,410,287,)
        self.line(5,30,415,30)
        self.line(35,5,35,30)
        self.line(385,5,385,30)
        self.rect(65,32,290,20)
        self.line(8,29,31,29)
        self.line(389,29,413,29)
        # self.line(5,47,292,47)
        # self.line(5,109,292,109)
        # self.line(5,171,292,171)
        self.rect(180,260,35,30)

    def texts(self):
        self.set_font('Arial',size = 10)
        self.set_xy(80,35)
        #self.cell(237,29,txt='Roll: {0}        Name: {1}               Year of admission: {2}'.format(rollno,name,yr))
        self.cell(237,10,txt= text_det , ln = 1, align = 'C'      )
        self.set_xy(130,45)
        self.cell(150,10,txt = text_det_s , ln = 2, align = 'C')
        #self.cell(237,10,txt = 'Programme : Bachelor of technology  Course: ', ln = 2, align = 'C')

    def text_interim(self,x,y):
        self.set_xy(x,y)
        self.set_font('Arial',size =6)
        self.cell(25,5,txt='INTERIM TRANSCRIPT', ln = 1, align = 'R')

    def date_time(self):
        self.set_xy(20,270)
        self.set_font('Arial',size = 15)
        self.cell(1,1,txt = 'Date and Time : {0}'.format(dt_string) , ln =1, align = 'L')
        self.line(21,273,58,273)
        # self.set_xy(380,280)
        # self.set_font('Arial',size =10)
        # self.cell(40,1,txt ='Assistant Registrar (Academic)' ,ln = 1 , align='L')
    def sign(self):
        self.set_xy(335,270)
        self.set_font('Arial',size = 15)
        self.cell(1,1,txt='Assistant Registrar (Academic)' ,ln = 1 , align='L')
        self.line(336,267,409,267)

    def gradelist(self):
        self.set_font('arial', 'B', 7)

        # sem 1
        self.set_xy(15,50)
        self.cell(10)
        self.cell(30, 10,'Semester1', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        self.cell(-10)
        heading_list = list(sheets_data['Sem1'].columns)
        print(heading_list)
        for head in header[:-1]:
          self.cell(15, 5, head, 1, 0, 'C')
          self.cell(15, 5, header[-1], 1, 2, 'C')
          self.cell(-60)
          self.set_font('arial', '', 6)
          for row in range(0, len(sheets_data['Sem1'])):
            for column_no, column_name in enumerate(header):
                if column_no != len(header) - 1:
                  self.cell(15, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                else:
                  self.cell(15, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                  self.cell(-60)
          self.cell(15, 10, "", 0, 2)
          self.cell(20)

output = 'output'
transcripts = 'transcriptsIITP'
if not os.path.exists(transcripts):
    os.makedirs(transcripts)

dir_path = os.path.join(os.getcwd(), transcripts)
folder_path = os.path.join(output)
# files = os.listdir(folder_path)
filenames = glob.glob(folder_path+"/*.xlsx")
for file in filenames:
  f_name = str(file)
  if all == 1:
    if f_name == fro_rol:
      for i in range(0,numb):
        #print(file)
        split__ = f_name.split('.')
        rol = str(split__[0])
        #print(rol)
        roll = rol[7:]
        yr = rol[7:9]
        year = '20'+yr
        #print(year)
        prog = rol[9:11]
        if prog == '01':
            programme = 'Bachelor of Technology'
        print(programme)
        cour = rol[11:13]
        if cour == 'ME':
            course = 'Mechanical engineering'
        elif cour == 'EE':
            course = 'Electrical Engineering'
        elif cour == 'CS':
            course = 'Computer Science Engineering'
        else:
            course = 'some other'
        #print(course)
        print('-----------\n')
        try:
          sheets_data = pd.read_excel(file,sheet_name = None )
          sheets = sheets_data.keys()

          #print(sheets_data['Sem1'])
          for row in sheets_data['Sem1']['Subject code']:
              print(row)
          sheet_keys = ['Sem1', 'Sem2', 'Sem3', 'Sem4', 'Sem5', 'Sem6', 'Sem7', 'Sem8']
          header = ['Subject code','Subject Name','L-T-P','Credit','Grade']
          pdf=PDF(orientation='L',unit='mm',format='A3')
          pdf.add_page()
          pdf.set_font('arial', 'B', 10)
          now = datetime.now()
          # dd/mm/YY H:M:S
          dt_string = now.strftime("%B %m,%Y   %H:%M:%S")
          grade_dic = {'BB': 8, 'BC': 7, 'AB' : 9, 'CC' : 6, 'AA' : 10, 'CD' : 5, 'DD' : 4, 'F' : 0, 'F*' : 0, 'DD*' : 4, ' BB' : 8}
          text_det = 'Roll number : {0}              Name : {1}                     Year of admission : {2} '.format(roll,names_roll[roll],year)   
          text_det_s = 'Programme : {0}    Course : {1}'.format(programme,course)

          pdf.border_lines()
          pdf.date_time()
          pdf.sign()
          pdf.texts()
          pdf.text_interim(7,25)
          pdf.text_interim(388,25)
          pdf.image('iitp_logo.png',x=8,y=5,w=25,h=20)
          pdf.image('iitp_logo.png',x=388,y=5,w=25,h=20)
          pdf.image('iitp-hin2.png',x=60,y=5,w=290,h=25)
          pdf.image(pic,x = 180, y= 260, w= 35,h = 30)

          # sem 1
          try:
            pdf.set_xy(12,50)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester1', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem1'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem1'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 2
          try:
            pdf.set_xy(110,50)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester2', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem2'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem2'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 3
          try:
            pdf.set_xy(208,50)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester3', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem3'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem3'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 4
          try:
            pdf.set_xy(306,50)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester4', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem4'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem4'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 5
          try:
            pdf.set_xy(12,130)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester5', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem5'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem5'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 6
          try:
            pdf.set_xy(110,130)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester6', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem6'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem6'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 7
          try:
            pdf.set_xy(208,130)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester7', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem7'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem7'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass

          # for sem 8
          try:
            pdf.set_xy(306,130)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester8', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem8'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem8'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass
          #sem9
          try:
            pdf.set_xy(12,210)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester9', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem9'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem9'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass
            #sem10
          try:
            pdf.set_xy(110,210)
            pdf.cell(10)
            pdf.set_font('arial', 'B', 10)
            pdf.cell(30, 10,'Semester10', 0, 2, 'C')
            #pdf.cell(90, 10, '', 0, 2, 'C')
            pdf.cell(-10)
            pdf.set_font('arial', '', 6)
            heading_list = list(sheets_data['Sem10'].columns)
            #print(heading_list)
            for head in header[:-1]:
              if head == 'Subject Name':
                pdf.cell(50,5,head,1,0,'C')
              elif head == 'Subject code':
                pdf.cell(15,5,head,1,0,'C')
              else:
                pdf.cell(10, 5, head, 1, 0, 'C')
            pdf.cell(10, 5, header[-1], 1, 2, 'C')
            pdf.cell(-85)
            pdf.set_font('arial', '', 6)
            for row in range(0, len(sheets_data['Sem10'])):
              for column_no, column_name in enumerate(header):
                if column_name == 'Subject Name':
                  if column_no != len(header) - 1:
                    pdf.cell(50, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(50, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                elif column_name == 'Subject code':
                  if column_no != len(header) - 1:
                    pdf.cell(15, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(15, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)
                else:
                  if column_no != len(header) - 1:
                    pdf.cell(10, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
                  else:
                    pdf.cell(10, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                    pdf.cell(-85)

            pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
            pdf.cell(20)
          except KeyError:
            pass
          #pdf.output(rol+'__.pdf','F')
          p = pdf.output(rol+'__.pdf','F')
          os.path.join(dir_path,p)
        except:
          print('Completed')

  elif all == 0:
    #print(file)
    split__ = f_name.split('.')
    rol = str(split__[0])
    #print(rol)
    roll = rol[7:]
    yr = rol[7:9]
    year = '20'+yr
    #print(year)
    prog = rol[9:11]
    if prog == '01':
        programme = 'Bachelor of Technology'
    print(programme)
    cour = rol[11:13]
    if cour == 'ME':
        course = 'Mechanical engineering'
    elif cour == 'EE':
        course = 'Electrical Engineering'
    elif cour == 'CS':
        course = 'Computer Science Engineering'
    else:
        course = 'some other'
    #print(course)
    print('-----------\n')
    try:
      sheets_data = pd.read_excel(file,sheet_name = None )
      sheets = sheets_data.keys()

      #print(sheets_data['Sem1'])
      for row in sheets_data['Sem1']['Subject code']:
          print(row)
      sheet_keys = ['Sem1', 'Sem2', 'Sem3', 'Sem4', 'Sem5', 'Sem6', 'Sem7', 'Sem8']
      header = ['Subject code','Subject Name','L-T-P','Credit','Grade']
      pdf=PDF(orientation='L',unit='mm',format='A3')
      pdf.add_page()
      pdf.set_font('arial', 'B', 10)
      now = datetime.now()
      # dd/mm/YY H:M:S
      dt_string = now.strftime("%B %m,%Y   %H:%M:%S")
      grade_dic = {'BB': 8, 'BC': 7, 'AB' : 9, 'CC' : 6, 'AA' : 10, 'CD' : 5, 'DD' : 4, 'F' : 0, 'F*' : 0, 'DD*' : 4, ' BB' : 8}
      text_det = 'Roll number : {0}              Name : {1}                     Year of admission : {2} '.format(roll,names_roll[roll],year)   
      text_det_s = 'Programme : {0}    Course : {1}'.format(programme,course)

      pdf.border_lines()
      pdf.date_time()
      pdf.sign()
      pdf.texts()
      pdf.text_interim(7,25)
      pdf.text_interim(388,25)
      pdf.image('iitp_logo.png',x=8,y=5,w=25,h=20)
      pdf.image('iitp_logo.png',x=388,y=5,w=25,h=20)
      pdf.image('iitp-hin2.png',x=60,y=5,w=290,h=25)
      pdf.image(pic,x = 180, y= 260, w= 35,h = 30)

      # sem 1
      try:
        pdf.set_xy(12,50)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester1', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem1'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem1'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 2
      try:
        pdf.set_xy(110,50)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester2', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem2'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem2'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem2']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem1']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 3
      try:
        pdf.set_xy(208,50)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester3', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem3'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem3'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem3']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 4
      try:
        pdf.set_xy(306,50)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester4', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem4'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem4'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem4']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 5
      try:
        pdf.set_xy(12,130)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester5', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem5'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem5'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem5']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 6
      try:
        pdf.set_xy(110,130)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester6', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem6'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem6'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem6']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 7
      try:
        pdf.set_xy(208,130)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester7', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem7'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem7'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem7']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass

      # for sem 8
      try:
        pdf.set_xy(306,130)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester8', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem8'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem8'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem8']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass
      #sem9
      try:
        pdf.set_xy(12,210)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester9', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem9'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem9'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem9']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass
        #sem10
      try:
        pdf.set_xy(110,210)
        pdf.cell(10)
        pdf.set_font('arial', 'B', 10)
        pdf.cell(30, 10,'Semester10', 0, 2, 'C')
        #pdf.cell(90, 10, '', 0, 2, 'C')
        pdf.cell(-10)
        pdf.set_font('arial', '', 6)
        heading_list = list(sheets_data['Sem10'].columns)
        #print(heading_list)
        for head in header[:-1]:
          if head == 'Subject Name':
            pdf.cell(50,5,head,1,0,'C')
          elif head == 'Subject code':
            pdf.cell(15,5,head,1,0,'C')
          else:
            pdf.cell(10, 5, head, 1, 0, 'C')
        pdf.cell(10, 5, header[-1], 1, 2, 'C')
        pdf.cell(-85)
        pdf.set_font('arial', '', 6)
        for row in range(0, len(sheets_data['Sem10'])):
          for column_no, column_name in enumerate(header):
            if column_name == 'Subject Name':
              if column_no != len(header) - 1:
                pdf.cell(50, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(50, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            elif column_name == 'Subject code':
              if column_no != len(header) - 1:
                pdf.cell(15, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(15, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)
            else:
              if column_no != len(header) - 1:
                pdf.cell(10, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 0, 'C')
              else:
                pdf.cell(10, 5, str(sheets_data['Sem10']['%s' % (column_name)].iloc[row]), 1, 2, 'C')
                pdf.cell(-85)

        pdf.cell(55, 5, "CPI :  SPI: CREDITS TAKEN :  CREDITS CLEARED:", 1, 2,'L')
        pdf.cell(20)
      except KeyError:
        pass
      #pdf.output(rol+'__.pdf','F')
      p = pdf.output(rol+'__.pdf','F')
      os.path.join(dir_path,p)
    except:
      print('Completed')

  else:
    print('No pdfs generated')